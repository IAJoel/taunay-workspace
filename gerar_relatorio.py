# -*- coding: utf-8 -*-
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

sys.stdout.reconfigure(encoding="utf-8")

dados = [
    ("0811324-98.2024.8.19.0036", "0811324-98.2024.8.19.0036", "SIM", "Memoriais para sessão de julgamento 12/05/2026"),
    ("0815064-08.2025.8.19.0205", "0815064-08.2025.8.19.0205", "SIM", "Memoriais para sessão de julgamento 12/05/2026"),
    ("202600376783", "0843171-33.2023.8.19.0205", "NÃO", "Memoriais para sessão de julgamento 13/05/2026"),
    ("202600376795", "0836346-35.2025.8.19.0001", "NÃO", "Memoriais para sessão de julgamento 13/05/2026"),
    ("202600376799", "0001075-40.2022.8.19.0021", "NÃO", "Memoriais para sessão de julgamento 13/05/2026"),
    ("0957065-80.2024.8.19.0001", "0957065-80.2024.8.19.0001", "SIM", "Memoriais de apelação cível para julgamento"),
    ("0820039-44.2023.8.19.0205", "0820039-44.2023.8.19.0205", "SIM", "Memoriais de apelação cível para julgamento"),
    ("0841018-94.2023.8.19.0021", "0841018-94.2023.8.19.0021", "SIM", "Manifestação em cumprimento de sentença"),
    ("0819999-28.2024.8.19.0205", "0819999-28.2024.8.19.0205", "SIM", "Manifestação em procedimento comum cível"),
    ("0800804-29.2022.8.19.0043", "0800804-29.2022.8.19.0043", "SIM", "Petição informando dados e requerendo mandado"),
    ("0807412-34.2023.8.19.0067", "0807412-34.2023.8.19.0067", "SIM", "Impugnação ao laudo pericial"),
    ("0905712-64.2025.8.19.0001", "0905712-64.2025.8.19.0001", "SIM", "Contestação ao pedido inicial"),
    ("0890840-15.2023.8.19.0001", "0890840-15.2023.8.19.0001", "SIM", "Impugnação ao laudo pericial"),
    ("0800118-31.2025.8.19.0205", "0800118-31.2025.8.19.0205", "SIM", "Impugnação ao laudo pericial"),
    ("0848148-67.2025.8.19.0021", "0848148-67.2025.8.19.0021", "SIM", "Protocolo de petição em procedimento comum cível"),
    ("0030664-14.2021.8.19.0021", "0030664-14.2021.8.19.0021", "SIM", "Laudo pericial do engenheiro eletricista"),
    ("0848149-52.2025.8.19.0021", "0848149-52.2025.8.19.0021", "SIM", "Protocolo de petição em procedimento comum cível"),
    ("0969205-49.2024.8.19.0001", "0969205-49.2024.8.19.0001", "SIM", "Petição com OP e OF em ação indenizatória"),
    ("0969205-49.2024.8.19.0001", "0969205-49.2024.8.19.0001", "SIM", "Petição com OP e OF em ação indenizatória"),
    ("3069906-13.2026.8.19.0001", "3069906-13.2026.8.19.0001", "SIM", "Protocolo de contestação em procedimento comum cível"),
    ("0829959-08.2024.8.19.0205", "0829959-08.2024.8.19.0205", "SIM", "Manifestação sobre prova documental da parte autora"),
    ("0840330-65.2023.8.19.0205", "0840330-65.2023.8.19.0205", "SIM", "Laudo pericial em procedimento comum cível"),
    ("202601731770", "0004672-46.2019.8.19.0207", "NÃO", "Petição juntando depósito de condenação e honorários"),
    ("202601731770", "0004672-46.2019.8.19.0207", "NÃO", "Petição juntando depósito de condenação e honorários"),
    ("0832444-79.2022.8.19.0001", "0832444-79.2022.8.19.0001", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0832444-79.2022.8.19.0001", "0832444-79.2022.8.19.0001", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0822560-25.2024.8.19.0205", "0822560-25.2024.8.19.0205", "SIM", "Manifestação sobre documentação da parte autora"),
    ("0018518-68.2021.8.19.0205", "0018518-68.2021.8.19.0205", "NÃO", "Petição juntando depósito e requerendo extinção execução"),
    ("0018518-68.2021.8.19.0205", "0018518-68.2021.8.19.0205", "NÃO", "Petição juntando depósito e requerendo extinção execução"),
    ("0801476-31.2025.8.19.0205", "0801476-31.2025.8.19.0205", "SIM", "Manifestação sobre petição de desocupação do imóvel"),
    ("0824906-76.2024.8.19.0001", "0824906-76.2024.8.19.0001", "SIM", "Recurso de apelação em ação de indenização"),
    ("0847958-07.2025.8.19.0021", "0847958-07.2025.8.19.0021", "SIM", "Contestação ao pedido inicial"),
    ("0809079-92.2024.8.19.0205", "0809079-92.2024.8.19.0205", "SIM", "Protocolo de petição em procedimento comum cível"),
    ("0820039-44.2023.8.19.0205", "0820039-44.2023.8.19.0205", "NÃO", "Memoriais enviados por email para sessão 12/05/2026"),
    ("0820039-44.2023.8.19.0205", "0820039-44.2023.8.19.0205", "SIM", "Memoriais para sessão de julgamento 12/05/2026"),
    ("0820039-44.2023.8.19.0205", "0820039-44.2023.8.19.0205", "NÃO", "Memoriais enviados por email para sessão 12/05/2026"),
    ("0001075-40.2022.8.19.0021", "0001075-40.2022.8.19.0021", "NÃO", "Memoriais enviados por email para sessão 13/05/2026"),
    ("0836346-35.2025.8.19.0001", "0836346-35.2025.8.19.0001", "NÃO", "Memoriais enviados por email para sessão 13/05/2026"),
    ("0843171-33.2023.8.19.0205", "0843171-33.2023.8.19.0205", "NÃO", "Memoriais enviados por email para sessão 13/05/2026"),
    ("0815064-08.2025.8.19.0205", "0815064-08.2025.8.19.0205", "NÃO", "Memoriais enviados por email para sessão 12/05/2026"),
    ("0135498-65.2020.8.19.0001", "0135498-65.2020.8.19.0001", "NÃO", "Manifestação discordando do valor dos honorários periciais"),
    ("0030956-67.2019.8.19.0021", "0030956-67.2019.8.19.0021", "NÃO", "Petição juntando depósito de condenação e honorários"),
    ("0811324-98.2024.8.19.0036", "0811324-98.2024.8.19.0036", "NÃO", "Memoriais enviados por email para sessão 12/05/2026"),
    ("0062723-94.2017.8.19.0021", "0062723-94.2017.8.19.0021", "NÃO", "Manifestação discordando dos cálculos do contador"),
    ("3005724-55.2026.8.19.0021", "3005724-55.2026.8.19.0021", "SIM", "Contestação ao pedido inicial"),
    ("0815561-60.2023.8.19.0021", "0815561-60.2023.8.19.0021", "SIM", "Impugnação ao laudo pericial"),
    ("0804717-82.2024.8.19.0064", "0804717-82.2024.8.19.0064", "SIM", "Protocolo de petição em procedimento comum cível"),
    ("0058315-21.2021.8.19.0021", "0058315-21.2021.8.19.0021", "NÃO", "Petição juntando documentação solicitada pelo perito"),
    ("0834109-70.2022.8.19.0021", "0834109-70.2022.8.19.0021", "SIM", "Petição juntando documentação solicitada pelo perito"),
    ("0852040-52.2023.8.19.0021", "0852040-52.2023.8.19.0021", "SIM", "Petição juntando documentação solicitada pelo perito"),
    ("0840353-10.2025.8.19.0021", "0840353-10.2025.8.19.0021", "SIM", "Contestação ao pedido inicial"),
    ("0313838-94.2021.8.19.0001", "0313838-94.2021.8.19.0001", "NÃO", "Petição juntando quesitos e protestando por suplementares"),
    ("0806088-16.2024.8.19.0021", "0806088-16.2024.8.19.0021", "SIM", "Protocolo de petição em procedimento comum cível"),
    ("0822294-76.2022.8.19.0021", "0822294-76.2022.8.19.0021", "SIM", "Protocolo de petição em procedimento comum cível"),
    ("0313721-16.2015.8.19.0001", "0313721-16.2015.8.19.0001", "NÃO", "Petição juntando quesitos e protestando por suplementares"),
    ("0813984-09.2025.8.19.0205", "0813984-09.2025.8.19.0205", "SIM", "Petição juntando quesitos e protestando por suplementares"),
    ("0863297-40.2024.8.19.0021", "0863297-40.2024.8.19.0021", "SIM", "Contestação ao pedido inicial"),
    ("0826627-33.2024.8.19.0205", "0826627-33.2024.8.19.0205", "NÃO", "Petição juntando quesitos e protestando por suplementares"),
    ("0810527-36.2025.8.19.0021", "0810527-36.2025.8.19.0021", "SIM", "Petição juntando quesitos e protestando por suplementares"),
    ("0813162-92.2022.8.19.0021", "0813162-92.2022.8.19.0021", "SIM", "Petição juntando quesitos e protestando por suplementares"),
    ("0821870-97.2023.8.19.0021", "0821870-97.2023.8.19.0021", "SIM", "Petição juntando quesitos e protestando por suplementares"),
    ("0810681-25.2023.8.19.0021", "0810681-25.2023.8.19.0021", "SIM", "Petição juntando quesitos e protestando por suplementares"),
    ("0002600-84.2020.8.19.0067", "0002600-84.2020.8.19.0067", "NÃO", "Recurso de apelação em ação revisional de faturamento"),
    ("0836714-86.2022.8.19.0021", "0836714-86.2022.8.19.0021", "SIM", "Impugnação ao laudo pericial"),
    ("0809583-92.2024.8.19.0207", "0809583-92.2024.8.19.0207", "SIM", "Cumprimento de sentença para cobrança de honorários"),
    ("0000100-83.2022.8.19.0064", "0000100-83.2022.8.19.0064", "NÃO", "Manifestação sobre esclarecimentos periciais"),
    ("0000100-83.2022.8.19.0064", "0000100-83.2022.8.19.0064", "NÃO", "Petição informando transação para encerrar litígio"),
    ("0038832-73.2019.8.19.0021", "0038832-73.2019.8.19.0021", "NÃO", "Petição requerendo expedição de mandado de pagamento"),
    ("0005094-82.2021.8.19.0067", "0005094-82.2021.8.19.0067", "NÃO", "Manifestação reiterando impugnação ao laudo pericial"),
    ("0053715-54.2021.8.19.0021", "0053715-54.2021.8.19.0021", "NÃO", "Manifestação reiterando impugnação ao laudo pericial"),
    ("0072148-87.2013.8.19.0021", "0072148-87.2013.8.19.0021", "NÃO", "Petição informando recolhimento de custas de mandado"),
    ("0803040-79.2024.8.19.0205", "0803040-79.2024.8.19.0205", "SIM", "Petição requerendo expedição de mandado de pagamento"),
    ("0803529-79.2023.8.19.0067", "0803529-79.2023.8.19.0067", "SIM", "Petição sobre mandado de pagamento não juntado"),
    ("0034260-45.2017.8.19.0021", "0034260-45.2017.8.19.0021", "NÃO", "Petição esclarecendo pagamento de honorários periciais"),
    ("0280348-57.2016.8.19.0001", "0280348-57.2016.8.19.0001", "NÃO", "Manifestação sobre excessos nos cálculos da parte autora"),
    ("0804018-84.2023.8.19.0207", "0804018-84.2023.8.19.0207", "SIM", "Petição referente ao evento 31 em apelação cível"),
    ("0811633-34.2023.8.19.0205", "0811633-34.2023.8.19.0205", "SIM", "Memoriais de apelação cível para julgamento"),
    ("202600381380", "0002726-13.2015.8.19.0067", "NÃO", "Memoriais para sessão de julgamento 14/05/2026"),
    ("202600381446", "0009803-05.2017.8.19.0067", "NÃO", "Memoriais para sessão de julgamento 20/05/2026"),
    ("202600381450", "0915871-03.2024.8.19.0001", "NÃO", "Memoriais para sessão de julgamento 12/05/2026"),
    ("202600381464", "0832490-67.2024.8.19.0205", "NÃO", "Memoriais para sessão de julgamento 13/05/2026"),
    ("202600381469", "0815581-85.2022.8.19.0021", "NÃO", "Memoriais para sessão de julgamento 20/05/2026"),
    ("202600381476", "0813343-55.2024.8.19.0205", "NÃO", "Memoriais para sessão de julgamento 13/05/2026"),
    ("202600381548", "0019885-63.2022.8.19.0021", "NÃO", "Memoriais para sessão de julgamento 20/05/2026"),
    ("202600381569", "0010754-64.2022.8.19.0021", "NÃO", "Memoriais para sessão de julgamento 20/05/2026"),
    ("0895567-80.2024.8.19.0001", "0895567-80.2024.8.19.0001", "SIM", "Petição informando pagamento de condenação"),
    ("0895567-80.2024.8.19.0001", "0895567-80.2024.8.19.0001", "SIM", "Petição comprovando cumprimento da obrigação de fazer"),
    ("0801929-60.2024.8.19.0205", "0801929-60.2024.8.19.0205", "SIM", "Manifestação reiterando posição sobre provas"),
    ("0805051-17.2025.8.19.0021", "0805051-17.2025.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0816078-31.2024.8.19.0021", "0816078-31.2024.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0826419-82.2025.8.19.0021", "0826419-82.2025.8.19.0021", "SIM", "Justificação de provas conforme despacho saneador"),
    ("0829491-82.2022.8.19.0021", "0829491-82.2022.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0833698-86.2024.8.19.0205", "0833698-86.2024.8.19.0205", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0837000-26.2024.8.19.0205", "0837000-26.2024.8.19.0205", "SIM", "Manifestação reiterando posição sobre provas"),
    ("0859629-27.2025.8.19.0021", "0859629-27.2025.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0819089-68.2024.8.19.0021", "0819089-68.2024.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0838522-25.2023.8.19.0205", "0838522-25.2023.8.19.0205", "SIM", "Manifestação sobre provas e regularidade do faturamento"),
    ("0812994-82.2024.8.19.0001", "0812994-82.2024.8.19.0001", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0815031-85.2025.8.19.0021", "0815031-85.2025.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0827407-40.2024.8.19.0021", "0827407-40.2024.8.19.0021", "SIM", "Manifestação ratificando posição sobre provas"),
    ("0831927-43.2024.8.19.0021", "0831927-43.2024.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0836685-65.2024.8.19.0021", "0836685-65.2024.8.19.0021", "SIM", "Justificação de provas conforme ato ordinatório"),
    ("0836762-74.2024.8.19.0021", "0836762-74.2024.8.19.0021", "SIM", "Petição informando juntada de prova documental suplementar"),
    ("0836822-13.2025.8.19.0021", "0836822-13.2025.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0850360-61.2025.8.19.0021", "0850360-61.2025.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0854954-55.2024.8.19.0021", "0854954-55.2024.8.19.0021", "SIM", "Justificação de provas conforme decisão saneadora"),
    ("0839082-30.2024.8.19.0205", "0839082-30.2024.8.19.0205", "SIM", "Petição juntando documentos sobre débitos de consumo"),
    ("0812065-49.2024.8.19.0001", "0812065-49.2024.8.19.0001", "SIM", "Manifestação sobre provas e regularidade do medidor"),
    ("0835279-77.2022.8.19.0021", "0835279-77.2022.8.19.0021", "SIM", "Justificação de provas sobre débito e inadimplência"),
    ("0905716-04.2025.8.19.0001", "0905716-04.2025.8.19.0001", "SIM", "Justificação de provas sobre leituras e faturamentos"),
    ("0800387-92.2025.8.19.0036", "0800387-92.2025.8.19.0036", "SIM", "Petição juntando depósito voluntário de condenação"),
    ("0815906-89.2024.8.19.0021", "0815906-89.2024.8.19.0021", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0815906-89.2024.8.19.0021", "0815906-89.2024.8.19.0021", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0811594-66.2025.8.19.0205", "0811594-66.2025.8.19.0205", "SIM", "Petição informando cumprimento tempestivo das obrigações"),
    ("0806013-38.2023.8.19.0206", "0806013-38.2023.8.19.0206", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0806013-38.2023.8.19.0206", "0806013-38.2023.8.19.0206", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0812125-56.2023.8.19.0001", "0812125-56.2023.8.19.0001", "SIM", "Petição informando cumprimento das obrigações impostas"),
    ("0010343-10.2021.8.19.0036", "0010343-10.2021.8.19.0036", "NÃO", "Petição juntando depósito de condenação e honorários"),
    ("0829968-63.2025.8.19.0001", "0829968-63.2025.8.19.0001", "SIM", "Petição juntando depósito de condenação"),
    ("0800523-37.2025.8.19.0021", "0800523-37.2025.8.19.0021", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0800523-37.2025.8.19.0021", "0800523-37.2025.8.19.0021", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0808113-02.2024.8.19.0021", "0808113-02.2024.8.19.0021", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0808113-02.2024.8.19.0021", "0808113-02.2024.8.19.0021", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0967208-65.2023.8.19.0001", "0967208-65.2023.8.19.0001", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0967208-65.2023.8.19.0001", "0967208-65.2023.8.19.0001", "SIM", "Petição juntando depósito de condenação e honorários"),
    ("0822733-87.2022.8.19.0021", "0822733-87.2022.8.19.0021", "SIM", "Petição requerendo levantamento de valores consignados"),
]

pasta_saida = os.path.join(os.path.dirname(__file__), "relatórios_análise de petições")
os.makedirs(pasta_saida, exist_ok=True)

hoje = date.today().isoformat()
caminho = os.path.join(pasta_saida, f"relatorio_{hoje}.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Protocolos"

cabecalhos = [
    "Número que está no protocolo da petição",
    "Número que está na petição",
    "Os números dos protocolos nos portais do PJE e Eproc correspondem aos números informados nas petições?",
    "Teor da petição",
]

azul_escuro = "1F3864"
branco = "FFFFFF"
verde = "C6EFCE"
verde_texto = "276221"
vermelho = "FFC7CE"
vermelho_texto = "9C0006"
cinza_claro = "F2F2F2"

borda = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, texto in enumerate(cabecalhos, 1):
    cel = ws.cell(row=1, column=col, value=texto)
    cel.font = Font(bold=True, color=branco, size=10)
    cel.fill = PatternFill("solid", fgColor=azul_escuro)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border = borda

ws.row_dimensions[1].height = 50

for i, (prot, num, corr, teor) in enumerate(dados, 2):
    fundo_linha = cinza_claro if i % 2 == 0 else "FFFFFF"

    for col, valor in enumerate([prot, num, corr, teor], 1):
        cel = ws.cell(row=i, column=col, value=valor)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = borda

        if col == 3:
            if valor == "SIM":
                cel.fill = PatternFill("solid", fgColor=verde)
                cel.font = Font(bold=True, color=verde_texto)
            else:
                cel.fill = PatternFill("solid", fgColor=vermelho)
                cel.font = Font(bold=True, color=vermelho_texto)
        else:
            cel.fill = PatternFill("solid", fgColor=fundo_linha)

    ws.row_dimensions[i].height = 30

larguras = [28, 28, 22, 55]
for col, larg in enumerate(larguras, 1):
    ws.column_dimensions[get_column_letter(col)].width = larg

wb.save(caminho)
print(f"Planilha salva em: {caminho}")

pasta_dados = os.path.join(os.path.dirname(__file__), "dados")
removidos = 0
for arquivo in os.listdir(pasta_dados):
    caminho_arq = os.path.join(pasta_dados, arquivo)
    if os.path.isfile(caminho_arq):
        os.remove(caminho_arq)
        removidos += 1
print(f"{removidos} arquivo(s) removido(s) da pasta dados.")
