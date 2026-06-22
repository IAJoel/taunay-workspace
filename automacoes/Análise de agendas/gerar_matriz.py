# -*- coding: utf-8 -*-
"""
gerar_matriz.py

Modos de uso:
  python gerar_matriz.py
      Lê os três arquivos de entrada, monta a MATRIZ, grava o Excel e um JSON
      com os pares pendentes de análise de equivalência.

  python gerar_matriz.py --update '{"2":"SIM","5":"NÃO",...}'
      Abre a MATRIZ do dia, preenche a coluna Equivalência e recalcula
      Inconsistências com base no JSON fornecido (chave = número da linha Excel).
"""
import sys
import os
import glob
import json
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

sys.stdout.reconfigure(encoding="utf-8")

PASTA = os.path.dirname(os.path.abspath(__file__))
HOJE  = date.today().isoformat()

# ── Estilos ───────────────────────────────────────────────────────────────────
COR_HEADER    = "1F3864"
COR_BRANCO    = "FFFFFF"
COR_VERDE     = "C6EFCE"
COR_VERDE_TXT = "276221"
COR_VERM      = "FFC7CE"
COR_VERM_TXT  = "9C0006"
COR_AMARELO   = "FFEB9C"
COR_AMAR_TXT  = "9C5700"
COR_CINZA     = "F2F2F2"

BORDA = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

CABECALHOS = [
    "Prazo Interno",
    "Número CNJ > Pasta > Pasta Tarefa",
    "Petição protocolada - Procv (GED)",
    "Número do protocolo",
    "O protocolo reproduz o número da petição?",
    "Conteúdo da petição",
    "Nome > Evento",
    "Equivalência",
    "Notas de Apoio – Advs",
    "Nome > Responsável",
    "Conteúdo",
    "Inconsistências",
]

COL_E = 5   # "O protocolo reproduz o número da petição?"
COL_H = 8   # "Equivalência"

# ── Helpers ───────────────────────────────────────────────────────────────────
EXCLUIR_DO_GLOB = ("MATRIZ", "Base analisada", "pares_analise")

def detectar(prefixo):
    """Retorna o arquivo mais recente que começa com 'prefixo'."""
    matches = [
        m for m in glob.glob(os.path.join(PASTA, f"{prefixo}*.xlsx"))
        if not any(ex in os.path.basename(m) for ex in EXCLUIR_DO_GLOB)
    ]
    if not matches:
        raise FileNotFoundError(
            f"Arquivo '{prefixo}*.xlsx' não encontrado em:\n  {PASTA}"
        )
    return sorted(matches)[-1]

def acumular(prefixo, aba, key_col, rename_cols=None):
    """Lê TODOS os arquivos matching, faz union e deduplica por key_col (última entrada prevalece)."""
    matches = sorted([
        m for m in glob.glob(os.path.join(PASTA, f"{prefixo}*.xlsx"))
        if not any(ex in os.path.basename(m) for ex in EXCLUIR_DO_GLOB)
    ])
    if not matches:
        raise FileNotFoundError(
            f"Nenhum arquivo '{prefixo}*.xlsx' encontrado em:\n  {PASTA}"
        )
    frames = []
    for m in matches:
        try:
            df = ler_excel(m, aba)
            if rename_cols:
                df = df.rename(columns=rename_cols)
            frames.append(df)
        except Exception as e:
            print(f"  [AVISO] Ignorando {os.path.basename(m)}: {e}")
    combined = pd.concat(frames, ignore_index=True)
    combined[key_col] = combined[key_col].astype(str).str.strip()
    return combined.drop_duplicates(subset=key_col, keep="last")

def ler_excel(caminho, aba):
    try:
        return pd.read_excel(caminho, sheet_name=aba)
    except Exception:
        return pd.read_excel(caminho, sheet_name=0)

def cor_celula(ws, row_i, col_i, valor, fundo):
    cel = ws.cell(row=row_i, column=col_i)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border = BORDA

    if col_i == COL_E:
        v = str(valor).strip().upper()
        if v == "SIM":
            cel.fill = PatternFill("solid", fgColor=COR_VERDE)
            cel.font = Font(bold=True, color=COR_VERDE_TXT)
        elif v == "NÃO":
            cel.fill = PatternFill("solid", fgColor=COR_VERM)
            cel.font = Font(bold=True, color=COR_VERM_TXT)
        else:
            cel.fill = PatternFill("solid", fgColor=fundo)
    elif col_i == COL_H:
        v = str(valor).strip().upper()
        if v == "SIM":
            cel.fill = PatternFill("solid", fgColor=COR_VERDE)
            cel.font = Font(bold=True, color=COR_VERDE_TXT)
        elif v == "NÃO":
            cel.fill = PatternFill("solid", fgColor=COR_VERM)
            cel.font = Font(bold=True, color=COR_VERM_TXT)
        elif "ANALISAR" in str(valor).upper():
            cel.fill = PatternFill("solid", fgColor=COR_AMARELO)
            cel.font = Font(bold=True, color=COR_AMAR_TXT)
        else:
            cel.fill = PatternFill("solid", fgColor=fundo)
    else:
        cel.fill = PatternFill("solid", fgColor=fundo)


# ══════════════════════════════════════════════════════════════════════════════
# MODO --update: atualiza Equivalência e Inconsistências no Excel já gerado
# ══════════════════════════════════════════════════════════════════════════════
def modo_update(equivalencias_json: str):
    equivalencias = json.loads(equivalencias_json)  # {"2": "SIM", "5": "NÃO", ...}

    caminho = os.path.join(PASTA, f"MATRIZ_{HOJE}.xlsx")
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"MATRIZ não encontrada: {caminho}")

    wb = openpyxl.load_workbook(caminho)
    ws = wb["MATRIZ"]

    for row_str, equiv in equivalencias.items():
        row_i = int(row_str)
        fundo = COR_CINZA if row_i % 2 == 0 else COR_BRANCO

        # Atualizar col H (Equivalência)
        ws.cell(row=row_i, column=COL_H).value = equiv
        cor_celula(ws, row_i, COL_H, equiv, fundo)

        # Recalcular col L (Inconsistências)
        prot_val = str(ws.cell(row=row_i, column=COL_E).value or "").strip().upper()
        teor_val = str(ws.cell(row=row_i, column=6).value or "").strip()
        evt_val  = str(ws.cell(row=row_i, column=7).value or "").strip()
        itens = []
        if equiv == "NÃO":
            itens.append(f'Conteúdo da petição ("{teor_val}") não corresponde ao evento ("{evt_val}")')
        if prot_val == "NÃO":
            itens.append("Número do protocolo não corresponde ao número da petição")
        incons = " | ".join(itens)
        ws.cell(row=row_i, column=12).value = incons
        fundo_incons = COR_CINZA if row_i % 2 == 0 else COR_BRANCO
        cel_incons = ws.cell(row=row_i, column=12)
        cel_incons.fill = PatternFill("solid", fgColor=fundo_incons)
        cel_incons.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel_incons.border = BORDA

    wb.save(caminho)

    # Resumo pós-update
    sim = sum(1 for v in equivalencias.values() if v == "SIM")
    nao = sum(1 for v in equivalencias.values() if v == "NÃO")
    print(f"MATRIZ atualizada: {caminho}")
    print(f"  Equivalência SIM: {sim}")
    print(f"  Equivalência NÃO: {nao}")

    # Contar inconsistências totais na planilha
    incons_total = 0
    for row_i in range(2, ws.max_row + 1):
        v = ws.cell(row=row_i, column=12).value
        if v and str(v).strip():
            incons_total += 1
    print(f"  Inconsistências totais: {incons_total}")


# ══════════════════════════════════════════════════════════════════════════════
# MODO padrão: ler os três arquivos e gerar a MATRIZ
# ══════════════════════════════════════════════════════════════════════════════
def modo_gerar():
    AP_CNJ    = "Número CNJ > Pasta > Pasta Tarefa"
    AP_PRAZO  = "Prazo Interno"
    AP_EVENTO = "Nome > Evento"
    AP_NOTAS  = "Notas de Apoio - Advs"
    AP_RESP   = "Nome > Responsável"
    AP_CONT   = "Conteúdo"
    GED_KEY   = "Número CNJ > Pasta"
    GED_TIPO  = "Tipo Documento"
    PET_KEY   = "Número que está na petição"
    PET_NUM   = "Número que está no protocolo da petição"
    PET_TEOR  = "Teor da petição"
    PET_RENAME = {"Os números dos processos informados nas petições": PET_KEY}

    # Aprovados: arquivo mais recente (também aceita "Alterado de realizado para finalizado")
    try:
        arq_ap = detectar("Aprovados em lote")
    except FileNotFoundError:
        arq_ap = detectar("Alterado de realizado para finalizado")

    # GED e Petições: acumular todos os arquivos históricos
    print(f"[1/4] Lendo arquivos...")
    print(f"  Aprovados: {os.path.basename(arq_ap)}")

    df_ap  = ler_excel(arq_ap, "Relatório")
    df_ged = acumular("Base GED", "Relatório", GED_KEY)
    df_pet = acumular("Análise das petições", "Planilha1", PET_KEY, rename_cols=PET_RENAME)

    n_ged = len(glob.glob(os.path.join(PASTA, "Base GED*.xlsx")))
    n_pet = len(glob.glob(os.path.join(PASTA, "Análise das petições*.xlsx")))
    print(f"  GED: {len(df_ged)} registros acumulados ({n_ged} arquivo(s))")
    print(f"  Petições: {len(df_pet)} registros acumulados ({n_pet} arquivo(s))")

    # Detectar coluna de protocolo (nome pode variar levemente)
    PET_PROT = next(
        (c for c in df_pet.columns if "protocolos nos portais" in c.lower()), None
    )

    print(f"[2/4] Montando MATRIZ...")

    df = df_ap[[AP_PRAZO, AP_CNJ, AP_EVENTO, AP_NOTAS, AP_RESP, AP_CONT]].copy()
    df[AP_PRAZO] = pd.to_datetime(df[AP_PRAZO], errors="coerce")
    df = df.sort_values(AP_PRAZO).reset_index(drop=True)
    df = df[df[AP_RESP].astype(str).str.strip() != "RAMON VINICIUS BAPTISTA DA SILVA"].reset_index(drop=True)

    # Merges via CNJ (df_ged e df_pet já vêm deduplicados de acumular())
    df[AP_CNJ] = df[AP_CNJ].astype(str).str.strip()

    df["Petição protocolada - Procv (GED)"]        = df[AP_CNJ].map(df_ged.set_index(GED_KEY)[GED_TIPO]).fillna("#N/D")
    df["Número do protocolo"]                       = df[AP_CNJ].map(df_pet.set_index(PET_KEY)[PET_NUM]).fillna("#N/D").astype(str)
    df["O protocolo reproduz o número da petição?"] = df[AP_CNJ].map(df_pet.set_index(PET_KEY)[PET_PROT]).fillna("#N/D") if PET_PROT else "#N/D"
    df["Conteúdo da petição"]                       = df[AP_CNJ].map(df_pet.set_index(PET_KEY)[PET_TEOR]).fillna("#N/D")

    df["Equivalência"]   = "ANALISAR INDIVIDUALMENTE"
    df["Inconsistências"] = ""

    # Pré-calcular inconsistências para os que já têm protocolo divergente
    for i, row in df.iterrows():
        prot_val = str(row["O protocolo reproduz o número da petição?"]).strip().upper()
        itens = []
        if prot_val == "NÃO":
            itens.append("Número do protocolo não corresponde ao número da petição")
        df.at[i, "Inconsistências"] = " | ".join(itens)

    # Reordenar colunas A–L
    df = df[[
        AP_PRAZO,
        AP_CNJ,
        "Petição protocolada - Procv (GED)",
        "Número do protocolo",
        "O protocolo reproduz o número da petição?",
        "Conteúdo da petição",
        AP_EVENTO,
        "Equivalência",
        AP_NOTAS,
        AP_RESP,
        AP_CONT,
        "Inconsistências",
    ]]

    # Pares para análise de equivalência (GED e Teor preenchidos)
    # índice Excel = índice df + 2 (linha 1 = cabeçalho)
    pares = []
    for i, row in df.iterrows():
        if row["Petição protocolada - Procv (GED)"] != "#N/D" and row["Conteúdo da petição"] != "#N/D":
            pares.append({
                "row_excel": i + 2,
                "cnj":       str(row[AP_CNJ]),
                "ged":       str(row["Petição protocolada - Procv (GED)"]),
                "teor":      str(row["Conteúdo da petição"]),
                "evento":    str(row[AP_EVENTO]),
            })

    print(f"[3/4] Gerando Excel ({len(df)} linhas, {len(pares)} pares para análise)...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MATRIZ"

    # Cabeçalho
    for col_i, texto in enumerate(CABECALHOS, 1):
        cel = ws.cell(row=1, column=col_i, value=texto)
        cel.font = Font(bold=True, color=COR_BRANCO, size=10)
        cel.fill = PatternFill("solid", fgColor=COR_HEADER)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA
    ws.row_dimensions[1].height = 45

    # Dados
    LARGURAS = [14, 32, 28, 22, 18, 35, 35, 14, 30, 30, 40, 40]
    for row_i, row in enumerate(df.itertuples(index=False), 2):
        fundo = COR_CINZA if row_i % 2 == 0 else COR_BRANCO
        for col_i, valor in enumerate(row, 1):
            if col_i == 1:
                try:
                    valor = pd.Timestamp(valor).date() if pd.notna(valor) else ""
                except Exception:
                    valor = "" if pd.isna(valor) else valor
            exibir = "" if (valor is None or (isinstance(valor, float) and pd.isna(valor))) else valor
            ws.cell(row=row_i, column=col_i).value = exibir
            cor_celula(ws, row_i, col_i, exibir, fundo)
        ws.row_dimensions[row_i].height = 30

    for col_i, larg in enumerate(LARGURAS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = larg

    saida_xlsx = os.path.join(PASTA, f"MATRIZ_{HOJE}.xlsx")
    wb.save(saida_xlsx)

    # JSON de pares
    saida_json = os.path.join(PASTA, f"pares_analise_{HOJE}.json")
    with open(saida_json, "w", encoding="utf-8") as f:
        json.dump(pares, f, ensure_ascii=False, indent=2)

    print(f"[4/4] Concluído.")
    print(f"\n  Excel:  MATRIZ_{HOJE}.xlsx")
    print(f"  JSON:   pares_analise_{HOJE}.json")
    print(f"  Linhas: {len(df)}")
    print(f"  Pares para análise de equivalência: {len(pares)}")


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--update", metavar="JSON", default=None,
                        help='JSON com equivalências: \'{"2":"SIM","5":"NÃO",...}\'')
    parser.add_argument("--update-file", metavar="FILE", default=None,
                        help='Arquivo JSON com equivalências')
    args = parser.parse_args()

    if args.update_file:
        with open(args.update_file, encoding="utf-8-sig") as f:
            modo_update(f.read().strip())
    elif args.update:
        modo_update(args.update)
    else:
        modo_gerar()
