# -*- coding: utf-8 -*-
"""
gerar_matriz_direto.py
Gera MATRIZ_SENTENCAS_YYYY-MM-DD.xlsx com os dados analisados diretamente.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

PASTA = os.path.dirname(os.path.abspath(__file__))
HOJE  = date.today()
DATA_SITUACAO = HOJE - timedelta(days=2)

COR_HEADER = "1F3864"
COR_BRANCO = "FFFFFF"
COR_CINZA  = "F2F2F2"

BORDA = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

CABECALHOS = [
    "Número do processo",
    "Fase atual",
    "1ª Inst. Dano Moral",
    "1ª Inst. Dano Material",
    "Situação 1ª instância",
    "Data situação 1ª inst.",
    "Juiz 1ª instância",
    "Quantidade Dias Sem Energia",
    "Matéria",
]

LARGURAS = [34, 20, 20, 22, 36, 22, 44, 28, 44]

# Dados analisados: (CNJ, dano_moral, dano_material, situacao, juiz, dias_sem_energia, fase_atual, materia)
# dano_moral: valor em centavos (ex: 300000 = R$3.000,00) ou "000" se improcedente
# dano_material: valor em centavos ou 0 se improcedente/ilíquido
# dias_sem_energia: string com número ou "" se não aplicável
# fase_atual: valor do campo "Nome > Evento" do arquivo de entrada
# materia: valor da coluna "Descrição dos Objetos > Pasta > Pasta Tarefa"
# Gerado em 25/06/2026 a partir de Follow Up Diario-25-06-2026_09_44_36.xlsx
DADOS = [
    # --- 01 | 0004541-42.2022.8.19.0021 — Belmiro Fontoura Ferreira Gonçalves
    ("0004541-42.2022.8.19.0021", "500000", "0", "Procedente",
     "Belmiro Fontoura Ferreira Gonçalves", "", "Sentença - Procedente em Parte",
     "COBRANÇA POR IRREGULARIDADE"),
    # --- 02 | 0016262-82.2018.8.19.0036 — Leandro Loyola de Abreu
    ("0016262-82.2018.8.19.0036", "000", "0", "Improcedente",
     "Leandro Loyola de Abreu", "", "Sentença - Improcedente",
     "FATURAS - ERRO DE LEITURA"),
    # --- 03 | 0077601-53.2019.8.19.0021 — sem identificação de juiz
    ("0077601-53.2019.8.19.0021", "500000", "0", "Procedente",
     "Não identificado", "", "Sentença - Procedente em Parte",
     "INTERRUPÇÃO DO FORNECIMENTO DE ENERGIA ELÉTRICA"),
    # --- 04 | 0081943-31.2023.8.19.0001 — Leandro Loyola de Abreu (34h → sem dias explícitos)
    ("0081943-31.2023.8.19.0001", "300000", "0", "Procedente",
     "Leandro Loyola de Abreu", "", "Sentença - Procedente em Parte",
     "INTERRUPÇÃO DO FORNECIMENTO DE ENERGIA ELÉTRICA"),
    # --- 05 | 0800517-98.2023.8.19.0021 — Livia Bechara de Castro
    ("0800517-98.2023.8.19.0021", "800000", "0", "Procedente",
     "Livia Bechara de Castro", "", "Sentença - Procedente em Parte",
     "Telemeditação – Corte/Suspensão"),
    # --- 06 | 0808060-44.2022.8.19.0036 — Priscila Abreu David
    ("0808060-44.2022.8.19.0036", "000", "0", "Improcedente",
     "Priscila Abreu David", "", "Sentença - Improcedente",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 07 | 0810220-15.2025.8.19.0205 — sem corpo de sentença (apenas dispositivo)
    ("0810220-15.2025.8.19.0205", "000", "0", "Improcedente",
     "Não identificado", "", "Sentença - Improcedente",
     "RELIGAÇÃO"),
    # --- 08 | 0815003-25.2022.8.19.0021 — Mirian Aninger Murad
    ("0815003-25.2022.8.19.0021", "000", "0", "Improcedente",
     "Mirian Aninger Murad", "", "Sentença - Improcedente",
     "CADASTRO - ENCERRAMENTO DE CONTRATO"),
    # --- 09 | 0821240-07.2024.8.19.0021 — Danilo Nunes Cronemberger Miranda
    ("0821240-07.2024.8.19.0021", "600000", "0", "Procedente",
     "Danilo Nunes Cronemberger Miranda", "", "Sentença - Procedente em Parte",
     "FATURAS - DIVERGÊNCIA QUANTO A ITENS FATURADOS"),
    # --- 10 | 0823127-52.2025.8.19.0001 — Luiz Claudio Silva Jardim Marinho
    ("0823127-52.2025.8.19.0001", "400000", "0", "Procedente",
     "Luiz Claudio Silva Jardim Marinho", "", "Sentença - Procedente em Parte",
     "COBRANÇA - NEGATIVAÇÃO"),
    # --- 11 | 0826509-27.2024.8.19.0021 — Danilo Nunes Cronemberger Miranda
    ("0826509-27.2024.8.19.0021", "000", "0", "Improcedente",
     "Danilo Nunes Cronemberger Miranda", "", "Sentença - Improcedente",
     "CADASTRO - ENCERRAMENTO DE CONTRATO"),
    # --- 12 | 0832194-83.2022.8.19.0021 — Maria Daniella Binato de Castro
    ("0832194-83.2022.8.19.0021", "500000", "0", "Procedente",
     "Maria Daniella Binato de Castro", "", "Sentença - Procedente em Parte",
     "PRAZOS - RELIGAÇÃO"),
    # --- 13 | 0833115-38.2023.8.19.0205 — sem corpo de sentença (apenas dispositivo)
    ("0833115-38.2023.8.19.0205", "000", "0", "Improcedente",
     "Não identificado", "", "Sentença - Improcedente",
     "CADASTRO - ABERTURA DE CONTRATO"),
    # --- 14 | 0835880-83.2022.8.19.0021 — Isabel Teresa Pinto Coelho Diniz (7 dias)
    ("0835880-83.2022.8.19.0021", "500000", "0", "Procedente",
     "Isabel Teresa Pinto Coelho Diniz", "7", "Sentença - Procedente em Parte",
     "INTERRUPÇÃO DO FORNECIMENTO DE ENERGIA ELÉTRICA"),
    # --- 15 | 0839330-97.2023.8.19.0021 — Isabel Teresa Pinto Coelho Diniz (dano material R$852,75)
    ("0839330-97.2023.8.19.0021", "600000", "85275", "Procedente",
     "Isabel Teresa Pinto Coelho Diniz", "", "Sentença - Procedente em Parte",
     "INTERRUPÇÃO DO FORNECIMENTO DE ENERGIA ELÉTRICA"),
    # --- 16 | 0919015-48.2025.8.19.0001 — Milena Angelica Drumond Morais Diz (só refaturamento)
    ("0919015-48.2025.8.19.0001", "000", "0", "Procedente",
     "Milena Angelica Drumond Morais Diz", "", "Sentença - Procedente em Parte",
     "FATURAS"),
    # --- 17 | 0961986-19.2023.8.19.0001 — Marcia Regina Sales Cardoso de Oliveira
    ("0961986-19.2023.8.19.0001", "700000", "0", "Procedente",
     "Marcia Regina Sales Cardoso de Oliveira", "", "Sentença - Procedente em Parte",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 18 | 0006722-05.2021.8.19.0036 — extinção por óbito (sem juiz identificado)
    ("0006722-05.2021.8.19.0036", "000", "0", "Extinto S/ Julg. Mérito – Demais Hipóteses",
     "Não identificado", "", "Sentença - Extinção sem julgamento de Mérito",
     "INTERRUPÇÃO DO FORNECIMENTO DE ENERGIA ELÉTRICA"),
    # --- 19 | 0011112-30.2020.8.19.0205 — ATENÇÃO: conteúdo é julgamento de embargos de declaração (art.1022 CPC), não sentença original
    ("0011112-30.2020.8.19.0205", "000", "0", "Improcedente",
     "Não identificado", "", "Sentença - Improcedente",
     "FATURAS - DIVERGÊNCIA QUANTO A ITENS FATURADOS"),
    # --- 20 | 0018053-63.2020.8.19.0021 — extinção por óbito
    ("0018053-63.2020.8.19.0021", "000", "0", "Extinto S/ Julg. Mérito – Demais Hipóteses",
     "Não identificado", "", "Sentença - Extinção sem julgamento de Mérito",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 21 | 0055582-53.2019.8.19.0021 — ATENÇÃO: apenas ato de reenvio à publicação; conteúdo da sentença indisponível no DJe
    ("0055582-53.2019.8.19.0021", "000", "0", "Procedente",
     "Não identificado", "", "Sentença - Procedente em Parte",
     "COBRANÇA POR IRREGULARIDADE"),
    # --- 22 | 0142306-47.2024.8.19.0001 — sem juiz identificado no DJe
    ("0142306-47.2024.8.19.0001", "000", "0", "Improcedente",
     "Não identificado", "", "Sentença - Improcedente",
     "SUSPENSÃO INDEVIDA"),
    # --- 23 | 0807356-09.2022.8.19.0205 — extinção (sem julgamento de mérito)
    ("0807356-09.2022.8.19.0205", "000", "0", "Extinto S/ Julg. Mérito – Demais Hipóteses",
     "Não identificado", "", "Sentença - Extinção sem julgamento de Mérito",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 24 | 0807648-87.2023.8.19.0001 — Sandro Lucio Barbosa Pitassi
    ("0807648-87.2023.8.19.0001", "500000", "0", "Procedente",
     "Sandro Lucio Barbosa Pitassi", "", "Sentença - Procedente em Parte",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 25 | 0815856-93.2024.8.19.0205 — Pedro Antonio de Oliveira Junior (7 dias: 21/03 a 28/03)
    ("0815856-93.2024.8.19.0205", "500000", "0", "Procedente",
     "Pedro Antonio de Oliveira Junior", "7", "Sentença - Procedente em Parte",
     "INTERRUPÇÃO DO FORNECIMENTO DE ENERGIA ELÉTRICA"),
    # --- 26 | 0818783-32.2024.8.19.0205 — Telmira de Barros Mondego (sem dano moral)
    ("0818783-32.2024.8.19.0205", "000", "0", "Procedente",
     "Telmira de Barros Mondego", "", "Sentença - Procedente em Parte",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 27 | 0827540-15.2024.8.19.0205 — Fabelisa Gomes Leal (refaturamento extinto; dano moral R$2.000)
    ("0827540-15.2024.8.19.0205", "200000", "0", "Procedente",
     "Fabelisa Gomes Leal", "", "Sentença - Procedente em Parte",
     "FATURAS - AUMENTO DE CONSUMO"),
    # --- 28 | 0836022-49.2024.8.19.0205 — extinção por óbito — Fabelisa Gomes Leal
    ("0836022-49.2024.8.19.0205", "000", "0", "Extinto S/ Julg. Mérito – Demais Hipóteses",
     "Fabelisa Gomes Leal", "", "Sentença - Extinção sem julgamento de Mérito",
     "Suspensão indevida"),
    # --- 29 | 0841831-20.2024.8.19.0205 — Roberta dos Santos Braga
    ("0841831-20.2024.8.19.0205", "500000", "0", "Procedente",
     "Roberta dos Santos Braga", "", "Sentença - Procedente em Parte",
     "Cobrança - Negativação SERASA/SPC"),
    # --- 30 | 0843236-95.2023.8.19.0021 — Belmiro Fontoura Ferreira Gonçalves (só declaratório; sem DM)
    ("0843236-95.2023.8.19.0021", "000", "0", "Procedente",
     "Belmiro Fontoura Ferreira Gonçalves", "", "Sentença - Procedente em Parte",
     "COBRANÇA POR IRREGULARIDADE"),
    # --- 31 | 0848611-74.2022.8.19.0001 — Marcia Regina Sales Cardoso de Oliveira (DM improcedente)
    ("0848611-74.2022.8.19.0001", "000", "0", "Procedente",
     "Marcia Regina Sales Cardoso de Oliveira", "", "Sentença - Procedente em Parte",
     "COBRANÇA POR IRREGULARIDADE"),
    # --- 32 | 0856823-53.2024.8.19.0021 — Monalisa Renata Artifon
    ("0856823-53.2024.8.19.0021", "500000", "0", "Procedente",
     "Monalisa Renata Artifon", "", "Sentença - Procedente em Parte",
     "FATURAS - AUMENTO DE CONSUMO"),
]


def gravar_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MATRIZ"

    # Cabeçalho
    for col_i, texto in enumerate(CABECALHOS, 1):
        cel = ws.cell(row=1, column=col_i, value=texto)
        cel.font = Font(bold=True, color=COR_BRANCO, size=10)
        cel.fill = PatternFill("solid", fgColor=COR_HEADER)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA
    ws.row_dimensions[1].height = 40

    # Dados
    for row_i, (cnj, dm, dmat, sit, juiz, dias, fase, materia) in enumerate(DADOS, 2):
        fundo = COR_CINZA if row_i % 2 == 0 else COR_BRANCO
        valores = [
            cnj,
            fase,
            dm,
            dmat,
            sit,
            DATA_SITUACAO,
            juiz,
            dias,
            materia,
        ]
        for col_i, valor in enumerate(valores, 1):
            cel = ws.cell(row=row_i, column=col_i, value=valor)
            cel.fill = PatternFill("solid", fgColor=fundo)
            cel.border = BORDA
            if col_i == 6 and isinstance(valor, date):
                cel.number_format = "DD/MM/YYYY"
                cel.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row_i].height = 28

    for col_i, larg in enumerate(LARGURAS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = larg

    nome = f"MATRIZ_SENTENCAS_{HOJE.isoformat()}.xlsx"
    saida = os.path.join(PASTA, nome)
    if os.path.exists(saida):
        try:
            os.rename(saida, saida)
        except PermissionError:
            nome = f"MATRIZ_SENTENCAS_{HOJE.isoformat()}_v2.xlsx"
            saida = os.path.join(PASTA, nome)
    wb.save(saida)
    return saida


if __name__ == "__main__":
    saida = gravar_excel()
    print(f"Planilha gerada: {saida}")
    print(f"Linhas de dados: {len(DADOS)}")
    print(f"Data situação: {DATA_SITUACAO.strftime('%d/%m/%Y')}")
